# Programa __main__
# Requisitos: Este programa cria 3 pastas num arquivo xlsx.
# Autor: Jessica Prass
# Versão: 1.0.0
# Dados: 01/09/2022


from openpyxl import load_workbook


def main():
    wb = load_workbook(filename='planilhas/orçamento.xlsx')

    for item in ['Receita', 'Despesa', 'Resultado']:
        wb.create_sheet(title=item)

    wb.save('planilhas/orçamento.xlsx')


if __name__ == "__main__":
    main()
